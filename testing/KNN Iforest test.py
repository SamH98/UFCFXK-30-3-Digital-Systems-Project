import pandas as pd
import pyshark
from openpyxl import Workbook
from sklearn.ensemble import IsolationForest
import time
import os
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import classification_report, confusion_matrix

# set device information to be captured from
device = "192.168.0.30"
captureinterface = 'Ethernet'

# set up workbooks
book = Workbook()
sheet = book.active
book1 = Workbook()
sheet1 = book1.active
book2 = Workbook()
sheet2 = book1.active
book3 = Workbook()
sheet3 = book1.active


# main should always be running on a loop at start up
def main():
    folders()
    capturetraining()
    conversations(r'C:\Users\samue\Documents\testing\ ' + device + '/traindata.xlsx', '1')
    capturelive()
    conversations(r'C:\Users\samue\Documents\testing\ ' + device + '/livedata.xlsx', '2')
    iforest()


# set up folder if device hasnt got any training data
def folders():
    newdevice = r'C:\Users\samue\Documents\testing\ ' + device  # checks to see if a device folder exists if not creates a new diterctory
    if not os.path.isdir(newdevice):
        os.makedirs(newdevice)
    book2.save(r'C:\Users\samue\Documents\testing\ ' + device + '/ldata.xlsx')
    book3.save(r'C:\Users\samue\Documents\testing\ ' + device + '/tdata.xlsx')


# run for 24 hours to capture training data
def capturetraining():
    book.save(r'C:\Users\samue\Documents\testing\ ' + device + '/traindata.xlsx')
    trainingdata = pd.DataFrame(columns=['Source IP', 'Destination IP', 'Protocol',
                                         'Size'], )  # data frame for holding captured packet information

    t_end = time.time() + 10  # 86400 #24hours run time

    capture = pyshark.LiveCapture(interface=captureinterface,
                                  display_filter="ip.addr == " + device)  # take in device information

    while time.time() < t_end:  # run for the 24 hour cycle
        capture.sniff(timeout=5)
        capture
        for packet in capture.sniff_continuously(packet_count=10):
            row = (packet.ip.src, packet.ip.dst, packet.highest_layer,
                   packet.captured_length)  # save packet infomration to data frame
            df_length = len(trainingdata)
            trainingdata.loc[df_length] = row

    trainingdata.to_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/traindata.xlsx', index=False)


# run for 15 minute cycles to capture live data
def capturelive():
    book1.save(r'C:\Users\samue\Documents\testing\ ' + device + '/livedata.xlsx')
    live = pd.DataFrame(columns=['Source IP', 'Destination IP', 'Protocol', 'Size'], )
    t_end = time.time() + 60  # 900

    capture = pyshark.LiveCapture(interface=captureinterface, display_filter="ip.addr == " + device)

    while time.time() < t_end:
        capture.sniff(timeout=5)
        capture
        for packet in capture.sniff_continuously(packet_count=10):
            row = (packet.ip.src, packet.ip.dst, packet.highest_layer, packet.captured_length)
            df_length = len(live)
            live.loc[df_length] = row

    live.to_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/livedata.xlsx', index=False)


# take in capture files and compress into individual conversations
def conversations(file, stage):
    completeData = pd.DataFrame(
        columns=['Monitored device', 'Destination/Source IP', 'Protocol', 'Conversation size'], ) #sets up data frames for new file and ip record
    ipdf = pd.DataFrame(columns=['IP ADDRESS', 'Protocol'])
    data = pd.read_excel(file)

    df_length = len(ipdf)   #add current device ip to record
    ipdf.loc[df_length] = device

    for index, row in data.iterrows():
        for i in row[['Source IP']]:        #loop through every row in the file
            for k in row[["Destination IP"]]:
                for j in row[['Protocol']]:
                    if i == device: #check to see if that cell contains the device ip
                        if not ((ipdf['IP ADDRESS'] == k) & (ipdf['Protocol'] == j)).any(): #check to see if that conversation has been recorded
                            ipdone = (k, j) #add conversation to record
                            df_length = len(ipdf)
                            ipdf.loc[df_length] = ipdone

                            size = data[data["Source IP"] == k][data["Protocol"] == j]['Size'].sum()    #get total size of conversation
                            size = size + data[data["Destination IP"] == k][data["Protocol"] == j]['Size'].sum()

                            conversation = (device, k, j, size)     #save conversation to new file
                            df_length = len(completeData)
                            completeData.loc[df_length] = conversation
                    else:
                        if not ((ipdf['IP ADDRESS'] == i) & (ipdf['Protocol'] == j)).any():
                            ipdone = (i, j)
                            df_length = len(ipdf)
                            ipdf.loc[df_length] = ipdone
                            size = data[data["Source IP"] == i][data["Protocol"] == j]['Size'].sum()
                            size = size + data[data["Destination IP"] == i][data["Protocol"] == j]['Size'].sum()

                            conversation = (device, i, j, size)
                            df_length = len(completeData)
                            completeData.loc[df_length] = conversation
    #depending on the stage save conversation data as training or live
    if stage == '1':
        completeData.to_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/tdata.xlsx', index=False)
    else:
        completeData.to_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/ldata.xlsx', index=False)


# analyses the conversation files
def iforest():
    tdf = pd.read_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/tdata.xlsx') #data frame of both train and live data
    tdf.head()
    ldf = pd.read_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/ldata.xlsx')
    ldf.head()

    for col in tdf.columns:     #change all non int objects in the data frame
        if tdf[col].dtype == "object":
            le = LabelEncoder()
            tdf[col].fillna("None", inplace=True)
            le.fit(list(tdf[col].astype(str).values))
            tdf[col] = le.transform(list(tdf[col].astype(str).values))

    for col in ldf.columns:
        if ldf[col].dtype == "object":
            le = LabelEncoder()
            ldf[col].fillna("None", inplace=True)
            le.fit(list(ldf[col].astype(str).values))
            ldf[col] = le.transform(list(ldf[col].astype(str).values))

    trainx = tdf.iloc[:, :-1].values
    trainy = tdf.iloc[:, 3].values

    livex = tdf.iloc[:, 3].values

    X_train, X_test, y_train, y_test = train_test_split(trainx, trainy, test_size=0.20) #split and scale the data
    scaler = StandardScaler()
    scaler.fit(X_train)
    X_train = scaler.transform(X_train)
    X_test = scaler.transform(X_test)

    livex_test = scaler.transform(livex)

    model = IsolationForest(n_estimators=50, max_samples='auto', contamination=float(0.1)) #set up the model and fit it for training

    model.fit(X_train, y_train)

    # trainingdata["iforest"].value_counts()
    # test["iforest"].value_counts()
    # df['scores']=model.decision_function(df)

    tdf['anomaly'] = pd.Series(model.predict(X_test))       #make a new column for if the data is annomaolous or not
    ldf['anomaly'] = pd.Series(model.predict(livex_test))

    ldf.to_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/results.xlsx', index=False)
    print(ldf.to_string(index=False))
    print(tdf.to_string(index=False))



def KNN():
    tdf = pd.read_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/tdata.xlsx')  # data frame of both train and live data
    tdf.head()
    ldf = pd.read_excel(r'C:\Users\samue\Documents\testing\ ' + device + '/ldata.xlsx')
    ldf.head()

    for col in tdf.columns:
        if tdf[col].dtype == "object":
            le = LabelEncoder()
            tdf[col].fillna("None", inplace=True)
            le.fit(list(tdf[col].astype(str).values))
            tdf[col] = le.transform(list(tdf[col].astype(str).values))

    for col in ldf.columns:
        if ldf[col].dtype == "object":
            le = LabelEncoder()
            ldf[col].fillna("None", inplace=True)
            le.fit(list(ldf[col].astype(str).values))
            ldf[col] = le.transform(list(ldf[col].astype(str).values))

            X = tdf.iloc[:, :-1].values
            y = tdf.iloc[:, 3].values
            Z = ldf.iloc[:, 3].values

            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.20)

            scaler = StandardScaler()
            scaler.fit(X_train)

            X_train = scaler.transform(X_train)
            X_test = scaler.transform(X_test)
            livex_test = scaler.transform(Z)

            classifier = KNeighborsClassifier(n_neighbors=5)
            classifier.fit(X_train, y_train)

            train_pred = classifier.predict(X_test)
            live_pred = classifier.predict(livex_test)



            print(confusion_matrix(y_test, train_pred))
            print(classification_report(y_test, train_pred))





main()
